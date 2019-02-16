def new_xlsx(File_Name):
    import openpyxl
    from openpyxl.styles import Font
    wb = openpyxl.Workbook()
    sheet=wb.active
    sheet.cell(row=1, column=1).font = Font(size=11, name='Calibri', bold=True)  # giving styling to each cell
    sheet.cell(row=1, column=2).font = Font(size=11, name='Calibri', bold=True)
    sheet.cell(row=1, column=1).value="Reassigned"
    sheet.cell(row=1, column=2).value="Not-Reassigned"
    wb.save(File_Name)


def updating_to_xlsx(File_Name,num,Flag,BugId,status_flag=None):
    import openpyxl
    wb=openpyxl.load_workbook(File_Name)
    sheet=wb.active
    if Flag ==1:
        if status_flag ==2:
            sheet.cell(row=(num - 1), column=2).value=""
            sheet.cell(row=num-1, column=Flag).value = BugId
            wb.save(File_Name)
            return (0)

        if (BugId != sheet.cell(row=(num-1),column=1).value):
            sheet.cell(row=num,column=Flag).value=BugId
            wb.save(File_Name)
            return (+1)

        else:
            #wb.save(File_Name)
            return (0)
    elif Flag == 2:
        if status_flag==1:
            sheet.cell(row=(num - 1), column=1).value = ""
            sheet.cell(row=num-1, column=Flag).value = BugId
            wb.save(File_Name)
            return (0)

        if (BugId != sheet.cell(row=(num-1),column=2).value):
            sheet.cell(row=num,column=Flag).value=BugId
            wb.save(File_Name)
            return (+1)

        else:
            #wb.save(File_Name)
            return (0)

    #wb.save(File_Name)